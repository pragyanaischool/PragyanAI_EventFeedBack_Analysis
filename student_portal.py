import streamlit as st
import datetime
from textblob import TextBlob
from openpyxl import Workbook, load_workbook
import os

# Database Path (Synced with app.py)
DB_FILE = "data/unified_knowledge_base.xlsx"

def save_entry(data):
    """Saves student feedback and profile to the Excel database."""
    if not os.path.exists("data"):
        os.makedirs("data")
        
    if not os.path.exists(DB_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "MasterData"
        ws.append(list(data.keys()))
        wb.save(DB_FILE)
        
    wb = load_workbook(DB_FILE)
    ws = wb.active
    ws.append(list(data.values()))
    wb.save(DB_FILE)

def render_student_flow():
    """Main rendering function for the Student Portal."""
    
    # --- LOGIN / REGISTRATION VIEW ---
    if not st.session_state.authenticated:
        st.markdown('<p class="main-header">Student Access</p>', unsafe_allow_html=True)
        col1, col2 = st.columns(2)
        
        with col1:
            u = st.text_input("Username")
            p = st.text_input("Password", type="password")
            if st.button("Access Portal"):
                if u and p: # Simplified Auth
                    st.session_state.authenticated = True
                    st.session_state.user = {"name": u, "role": "student", "usn": "1GN21CS001"}
                    st.rerun()
                else:
                    st.error("Please enter credentials.")
        with col2:
            st.info("PragyanAI ensures your voice is heard directly by the trainers. Your feedback helps us improve session quality.")
        return

    # --- DASHBOARD VIEW (LOGGED IN) ---
    st.markdown(f'<p class="main-header">Hello, {st.session_state.user["name"]}</p>', unsafe_allow_html=True)
    
    # Event Selection
    events = {
        "Multi-Agent AI Systems": {"trainer": "Sateesh A.", "topics": "RAG, AutoGen, CrewAI"},
        "Edge Intelligence": {"trainer": "Priya K.", "topics": "ESP32, MicroPython, MQTT"},
        "PCB Automation": {"trainer": "Dr. Ramesh", "topics": "KiCad, Schematic Design, Auto-routing"}
    }
    
    selected_event_name = st.selectbox("Select the Event you attended", list(events.keys()))
    evt = events[selected_event_name]
    
    st.markdown(f"""
    <div class="metric-card">
        <h4>📍 {selected_event_name}</h4>
        <p><b>Trainer:</b> {evt['trainer']} | <b>Core Focus:</b> {evt['topics']}</p>
    </div>
    """, unsafe_allow_html=True)

    st.divider()

    # --- FEEDBACK STAGES ---
    
    # Stage 1: Ratings & Topics
    if st.session_state.step == 1:
        with st.form("phase_1_form"):
            st.subheader("Step 1: Quantitative Evaluation")
            c1, c2, c3 = st.columns(3)
            r1 = c1.feedback("stars", key="rate_overall")
            r2 = c2.feedback("stars", key="rate_content")
            r3 = c3.feedback("stars", key="rate_hands_on")
            
            topics_covered = st.text_area("List specific topics you mastered today:")
            liked_most = st.text_input("What was the highlight of this session?")
            
            if st.form_submit_button("Proceed to Deep Analysis"):
                if r1 is None or not topics_covered:
                    st.warning("Please provide an overall rating and topics covered.")
                else:
                    st.session_state.temp_feedback = {
                        "r1": r1 + 1, "r2": (r2 or 0) + 1, "r3": (r3 or 0) + 1,
                        "topics": topics_covered, "liked": liked_most, 
                        "event": selected_event_name, "trainer": evt['trainer']
                    }
                    st.session_state.step = 2
                    st.rerun()
    
    # Stage 2: Media, Voice & Deep Text
    else:
        st.subheader("Step 2: Deep Feedback & Proof of Learning")
        if st.button("← Back to Step 1"):
            st.session_state.step = 1
            st.rerun()
            
        img = st.camera_input("Visual Validation (Snapshot of your work or session)")
        audio = st.audio_input("Record Voice Feedback (Explain your experience in 30 seconds)")
        detailed_txt = st.text_area("Detailed feedback/thoughts for deep analysis")
        
        if st.button("Submit Final Intelligence Report"):
            # Sentiment Analysis
            blob = TextBlob(detailed_txt)
            polarity = blob.sentiment.polarity
            sentiment_label = "Positive" if polarity > 0.1 else "Neutral" if polarity >= -0.1 else "Negative"
            
            final_data = {
                "Timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Username": st.session_state.user['name'],
                "USN": st.session_state.user['usn'],
                "Event": st.session_state.temp_feedback['event'],
                "Trainer": st.session_state.temp_feedback['trainer'],
                "Rating_Overall": st.session_state.temp_feedback['r1'],
                "Rating_Content": st.session_state.temp_feedback['r2'],
                "Rating_HandsOn": st.session_state.temp_feedback['r3'],
                "Topics": st.session_state.temp_feedback['topics'],
                "Liked": st.session_state.temp_feedback['liked'],
                "Sentiment": sentiment_label,
                "Detailed_Transcript": detailed_txt,
                "Media_Captured": "Yes" if img else "No"
            }
            
            save_entry(final_data)
            st.success("✅ Feedback successfully synced with Admin Intelligence Suite!")
            
            # Smart Redirects (Google Reviews)
            if final_data['Rating_Overall'] >= 4:
                st.balloons()
                st.markdown("### 🌟 You had an excellent session!")
                st.info("Since you enjoyed the workshop, help us grow by leaving a Google Review.")
                st.link_button("Review on Google", "https://g.page/r/your-google-id/review")
            
            # Reset Flow
            st.session_state.step = 1
            if st.button("Submit Another Review"):
                st.rerun()
              
