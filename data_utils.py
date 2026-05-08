import os
import json
import pandas as pd
from openpyxl import Workbook, load_workbook

# Global file paths
DB_FILE = "data/unified_knowledge_base.xlsx"
EVENTS_FILE = "data/events.json"

def init_folders():
    """Ensures necessary data directories exist on system startup."""
    if not os.path.exists("data"):
        os.makedirs("data")

def load_events():
    """Loads event configuration from the JSON database."""
    init_folders()
    if os.path.exists(EVENTS_FILE):
        try:
            with open(EVENTS_FILE, "r") as f:
                return json.load(f)
        except Exception:
            # Return empty dict if JSON is corrupted or unreadable
            return {}
    return {}

def save_events(events_dict):
    """Saves updated event configuration to the JSON database."""
    init_folders()
    try:
        with open(EVENTS_FILE, "w") as f:
            json.dump(events_dict, f, indent=4)
    except Exception as e:
        print(f"Error saving events: {e}")

def save_feedback_entry(data):
    """Saves a single feedback record to the master Excel knowledge base."""
    init_folders()
    
    # Create file with headers if it doesn't exist
    if not os.path.exists(DB_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "MasterData"
        ws.append(list(data.keys()))
        wb.save(DB_FILE)
    
    try:
        wb = load_workbook(DB_FILE)
        ws = wb.active
        # Append the values from the dictionary in the order of columns
        ws.append(list(data.values()))
        wb.save(DB_FILE)
    except Exception as e:
        print(f"Error saving feedback entry: {e}")

def get_feedback_df():
    """Returns the entire feedback database as a Pandas DataFrame for analysis."""
    if os.path.exists(DB_FILE):
        try:
            return pd.read_excel(DB_FILE)
        except Exception:
            return pd.DataFrame()
    return pd.DataFrame()

def export_to_json():
    """Converts the Excel database to a JSON dump for backend processing."""
    df = get_feedback_df()
    if not df.empty:
        return df.to_json(orient="records", indent=4)
    return "[]"
    
